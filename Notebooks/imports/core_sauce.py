# Notebooks/imports/core_sauce.py
import asyncio
import sys
import json
import pandas as pd
from pathlib import Path
from loguru import logger
from pipulate import wand
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

async def universal_scrape(job, urls, headless=True, delay_range=(5, 10)):
    """The unified acquisition engine for all tier-2 tools."""
    extracted_data = []
    for i, url in enumerate(urls):
        current_delay = delay_range if i > 0 else None
        result = await wand.scrape(url=url, headless=headless, delay_range=current_delay)
        
        is_cached = result.get("cached", False)
        status = "✅ Cached" if is_cached else "👁️ Scraped"
        print(f"  -> {status} [{i+1}/{len(urls)}] {url}")

        dom_path = result.get("looking_at_files", {}).get("rendered_dom")
        if dom_path and Path(dom_path).exists():
            with open(dom_path, 'r', encoding='utf-8') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                extracted_data.append({
                    'url': url,
                    'title': soup.title.string.strip() if soup.title else "No Title",
                    'h1s': [h1.get_text(strip=True) for h1 in soup.find_all('h1')],
                    'rendered_dom_path': str(dom_path)
                })
    wand.set(job, "extracted_data", extracted_data)
    return extracted_data

async def generate_optics_batch(job, verbose=False):
    """Fires the LLM Optics engine for every URL in the job state."""
    extracted_data = wand.get(job, "extracted_data", [])
    script_path = (Path(wand.paths.root) / "tools" / "llm_optics.py").resolve()
    
    tasks = []
    for item in extracted_data:
        dom_path = item.get('rendered_dom_path')
        if dom_path and Path(dom_path).exists():
            tasks.append(_run_optics_subprocess(sys.executable, script_path, dom_path))
            
    if tasks:
        logger.info(f"🚀 Launching {len(tasks)} LLM Optics subprocesses...")
        await asyncio.gather(*tasks)

async def _run_optics_subprocess(py_exe, script, dom):
    proc = await asyncio.create_subprocess_exec(
        py_exe, str(script), str(dom),
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE
    )
    await proc.communicate()
    return proc.returncode == 0

def format_excel_pro(job, df, sheet_name="Analysis"):
    """The master 'Painterly Pass' for Excel output."""
    output_dir = Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)
    xl_file = output_dir / f"{job}_output.xlsx"
    
    with pd.ExcelWriter(xl_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        # Add Table styling and Banded rows
        tab = Table(displayName="DataTable", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
        # Auto-width and Alignment
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 30
            for cell in col: cell.alignment = Alignment(wrap_text=True, vertical='top')
    print(f"✅ Success! Deliverable saved to: {xl_file}")
    return xl_file