# Notebooks/imports/onboard_sauce.py
# Purpose: Bridge the 0nboard notebook to the LLM Optics engine.
# Asserting sovereignty over perception. 👁️

import asyncio
import sys
import json
import socket
import urllib.request
import os
from pathlib import Path
import ipywidgets as widgets
from IPython.display import display
from loguru import logger
from pipulate import wand  # Use wand!
import imports
import llm
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


def check_ai_models(preferred_local=None, preferred_cloud=None):
    """
    Uses the Universal Adapter (llm) to verify AI readiness using fuzzy matching
    against a prioritized list of preferred models.
    """
    if preferred_local:
        wand.speak(f"Scanning for your preferred local models...")
    else:
        wand.speak("Scanning your system for available AI models...")

    try:
        # 1. Gather all models known to the Universal Adapter
        available_models = [m.model_id for m in llm.get_models()]
        
        # 2. Check for ANY local model (Ollama models typically lack provider prefixes)
        has_local = any('ollama' in str(type(m)).lower() for m in llm.get_models())

        # 3. Process User Preferences
        def parse_preferences(pref_string):
            if not pref_string: return []
            return [p.strip().lower() for p in pref_string.split(',')]

        local_prefs = parse_preferences(preferred_local)
        cloud_prefs = parse_preferences(preferred_cloud)

        selected_local = None
        selected_cloud = None

        # 4. Fuzzy Matching Logic (Find highest priority match)
        # We check each preference against the available models. If the preference
        # string is *in* the available model string (e.g., 'gemma3' in 'gemma3:latest'), it's a match.
        for pref in local_prefs:
            match = next((m for m in available_models if pref in m.lower() and 'ollama' in str(type(llm.get_model(m))).lower()), None)
            if match:
                selected_local = match
                break # Found our highest priority local model

        for pref in cloud_prefs:
            match = next((m for m in available_models if pref in m.lower() and 'ollama' not in str(type(llm.get_model(m))).lower()), None)
            if match:
                selected_cloud = match
                break # Found our highest priority cloud model

        # 5. Reporting and Graceful Degradation
        if selected_local:
            wand.speak(f"Excellent. Local model '{selected_local}' is active and ready.")
            print(f"\n✅ Locked in Local Model: {selected_local}")
        elif has_local:
            # Fallback: They have Ollama, but not their preferred model
            wand.speak("I found local models, but not your preferred choices.")
            print(f"\nℹ️  Preferred local models not found, but other local models are available.")
            print(f"Available models: {', '.join([m for m in available_models if 'ollama' in str(type(llm.get_model(m))).lower()])}")
            selected_local = True # Indicate local capacity exists
        else:
            # The Fallback State: No local models detected
            wand.speak("I do not detect a local AI brain on your system.")
            print("\nℹ️  Ollama is not running or not installed.")
            print("Pipulate works perfectly fine without it, but a local AI 'riding shotgun' ensures privacy.")
            print("\nTo upgrade your environment for true Local-First Sovereignty:")
            print("1. Go to https://ollama.com/")
            print("2. Download the installer for your host operating system.")
            print("3. Install it, open a terminal, run 'ollama run gemma3', and try again.")
            
        if selected_cloud:
             print(f"✅ Locked in Cloud Model: {selected_cloud}")

        return {
            "local": selected_local,
            "cloud": selected_cloud,
            "has_any_local": has_local
        }

    except Exception as e:
        print(f"❌ Error communicating with the Universal Adapter: {e}")
        return {"local": False, "cloud": False, "has_any_local": False}


def interrogate_local_ai(target_url: str, preferred_model: str = None):
    """Reads the accessibility tree and asks the local AI to summarize it."""
    from tools.scraper_tools import get_safe_path_component
    
    domain, url_path_slug = get_safe_path_component(target_url)

    md_file = wand.paths.browser_cache / domain / url_path_slug / "accessibility_tree.json"

    if md_file.exists():
        content = md_file.read_text()
        
        # Use first 2000 characters to keep it fast
        prompt = f"Based on the following DevTools accessibility tree extracted from a scrape, what is this page about? Answer in exactly 3 short bullet points.\n\n{content[:2000]}"
        
        try:
            # The Universal Adapter handles fallbacks automatically!
            if preferred_model:
                model = llm.get_model(preferred_model)
            else:
                model = llm.get_model()  # Auto-grabs the default
                
            target_model_id = model.model_id
            wand.speak(f"I am now interrogating the scraped data using the Universal Adapter, routed to {target_model_id}.")
            
            # The elegant prompt execution
            response = model.prompt(prompt)
            
            print(f"🤖 Analysis from {target_model_id}:\n")
            print(response.text())                
            wand.speak("Analysis complete. As you can see, I can read and summarize local files instantly.")

        except Exception as e:
            print(f"⚠️ Could not complete local AI analysis: {e}")
    else:
        print(f"⚠️ Could not find {md_file}. Did the previous step complete successfully?")


async def analyze_ai_readiness(job: str, url: str, verbose: bool = True, override_cache: bool = False):
    """
    The master 'Aha!' sequence for onboarding.
    Scrapes a URL and immediately shatters it into LLM Optics.
    """
    wand.speak(f"Beginning AI-Readiness analysis for {url}.")

    if override_cache:
        print("🧹 Cache override engaged. Forcing a fresh scrape.")
    
    # 1. THE SCRAPE (The Acquisition)
    if not override_cache:
        logger.info(f"🚀 Step 1: Checking cache or navigating to {url}...")

    result = await wand.scrape(
        url=url, 
        take_screenshot=True, 
        headless=False, 
        override_cache=override_cache,
        verbose=verbose
    )
    
    if not result.get('success'):
        error_msg = result.get('error', 'Unknown error')
        wand.speak("I encountered an issue during navigation.")
        print(f"❌ Scrape Failed: {error_msg}")
        return False

    if result.get('cached'):
        wand.speak("I already have this data cached locally. Bypassing browser navigation.")
        print("⚡ Cache Hit! Using existing artifacts to save time and compute.")
    else:
        wand.speak("Navigation complete. Page data captured.")
        print("✅ Fresh Scrape Successful.")

    # 2. THE OPTICS (The Refraction)
    if not dom_path or not Path(dom_path).exists():
        print("❌ Error: Could not locate hydrated_dom.html for analysis.")
        return False

    wand.speak("I have captured the page. Now, generating AI Optics.")
    logger.info(f"👁️‍🗨️ Step 2: Running LLM Optics Engine on {dom_path}...")
    
    optics_result = await generate_optics_subprocess(dom_path)
    
    if optics_result.get('success'):
        wand.speak("Analysis complete. You can now see your site through the eyes of an AI.")
        print(f"✅ Success! Optics generated in: {Path(dom_path).parent}")
        return True
    else:
        print(f"⚠️ Optics generation partially failed: {optics_result.get('error')}")
        return False

async def generate_optics_subprocess(dom_file_path: str):
    """Isolated wrapper to call llm_optics.py as a subprocess."""
    script_path = (Path(__file__).resolve().parent.parent.parent / "tools" / "llm_optics.py")
    
    proc = await asyncio.create_subprocess_exec(
        sys.executable, str(script_path), str(dom_file_path),
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE
    )
    
    stdout, stderr = await proc.communicate()
    
    if proc.returncode == 0:
        return {"success": True, "output": stdout.decode()}
    else:
        return {"success": False, "error": stderr.decode()}


def explain_optics_artifacts(target_url):
    """
    Feeds the AI both the raw source and the rendered DOM structure
    to educate the user on 'The JavaScript Gap'.
    """
    domain, slug = get_safe_path_component(target_url)
    base = wand.paths.browser_cache / domain / slug
    
    source_txt = (base / "source.html").read_text()[:1000] # Snippet
    rendered_txt = (base / "hydrated_dom.html").read_text()[:1000] # Snippet

    prompt = f"""
    I am looking at the scrape artifacts for {target_url}.
    
    RAW SOURCE SNIPPET:
    {source_txt}
    
    RENDERED DOM SNIPPET:
    {rendered_txt}
    
    Explain to the human why these two views differ and why a 'Forever Machine' 
    needs to see the Rendered DOM to understand modern web apps. 
    Point out any specific SEO tags or content that only appear in one of them.
    """
    
    response = wand.prompt(prompt)
    display(Markdown(f"### 👁️ Optics Education\n\n{response}"))


def ensure_cloud_credentials(cloud_model_id):
    """
    The Gatekeeper. Checks for required cloud API keys upfront to prevent 
    mid-workflow lazy-loading crashes. Renders a secure widget if missing.
    """
    import os
    import ipywidgets as widgets
    from IPython.display import display, clear_output
    from dotenv import load_dotenv, set_key

    # Load existing environment variables
    load_dotenv()
    
    if not cloud_model_id:
        print("ℹ️ No cloud model selected or required.")
        return

    # Map the model string to the standard API Key variable
    env_var_name = None
    if 'claude' in cloud_model_id.lower() or 'anthropic' in cloud_model_id.lower():
        env_var_name = 'ANTHROPIC_API_KEY'
    elif 'gpt' in cloud_model_id.lower() or 'openai' in cloud_model_id.lower():
        env_var_name = 'OPENAI_API_KEY'
    elif 'gemini' in cloud_model_id.lower() or 'google' in cloud_model_id.lower():
        env_var_name = 'GEMINI_API_KEY'
        
    if env_var_name and not os.getenv(env_var_name):
        wand.speak(f"I need your API key for {cloud_model_id} to proceed.")
        
        # Build the secure password widget (hides the input)
        key_input = widgets.Password(
            value='',
            placeholder='Paste your API Key here...',
            description=f'🔑 {env_var_name}:',
            style={'description_width': 'initial'},
            disabled=False,
            layout=widgets.Layout(width='80%')
        )
        
        submit_btn = widgets.Button(
            description="Save to Vault", 
            button_style='success',
            icon='lock'
        )
        out = widgets.Output()
        
        def on_submit(b):
            with out:
                clear_output()
                if key_input.value.strip():
                    # 1. Save permanently to .env
                    env_path = Path('.env')
                    env_path.touch(exist_ok=True)
                    set_key(str(env_path), env_var_name, key_input.value.strip())
                    
                    # 2. Export to current runtime environment
                    os.environ[env_var_name] = key_input.value.strip()
                    
                    # 3. Explicitly set it in Simon Willison's 'llm' tool keychain
                    try:
                        import llm
                        # The llm library uses specific aliases for its core plugins
                        if env_var_name == 'GEMINI_API_KEY':
                            key_alias = 'gemini'
                        elif env_var_name == 'OPENAI_API_KEY':
                            key_alias = 'openai'
                        elif env_var_name == 'ANTHROPIC_API_KEY':
                            key_alias = 'anthropic'
                        else:
                            key_alias = env_var_name.split('_')[0].lower()

                        llm.set_key(key_alias, key_input.value.strip())
                    except Exception as e:
                        print(f"⚠️ Warning: Could not sync key to internal LLM keychain: {e}")
                        
                    wand.speak("Key securely saved to the vault. The cloud is connected.")
                    wand.speak("You may now run the next cell.", emoji="✅") # REPLACED
                    wand.imperio()
                else:
                    print("❌ Please enter a valid API key.")
        
        submit_btn.on_click(on_submit)
        display(widgets.VBox([key_input, submit_btn, out]))
        
    else:
        # Key is already present, keep the rhythm flowing
        wand.speak("Cloud credentials verified in your environment.")
        print(f"✅ Secure connection ready for {cloud_model_id}.")


def etl_optics_to_excel(job: str, target_url: str):
    """
    The ETL (Extract, Transform, Load) demonstration.
    Extracts data from raw scraped artifacts (JSON, Markdown, TXT), 
    Transforms it into Pandas DataFrames, and Loads it into an Excel deliverable
    with conditional formatting to emulate a terminal output.
    """
    import pandas as pd
    import re
    import yaml
    import json
    import ipywidgets as widgets
    from tools.scraper_tools import get_safe_path_component
    from pipulate import wand

    domain, slug = get_safe_path_component(target_url)
    cache_dir = wand.paths.browser_cache / domain / slug

    # --- EXTRACT & TRANSFORM: SEO Metadata ---
    seo_file = cache_dir / "seo.md"
    seo_data = {"Metric": [], "Value": []}
    if seo_file.exists():
        content = seo_file.read_text(encoding='utf-8')
        match = re.search(r'^---\n(.*?)\n---', content, re.DOTALL)
        if match:
            try:
                frontmatter = yaml.safe_load(match.group(1))
                for k, v in frontmatter.items():
                    seo_data["Metric"].append(str(k).replace('_', ' ').title())
                    seo_data["Value"].append(str(v))
            except Exception as e:
                print(f"⚠️ Warning: Could not parse SEO frontmatter: {e}")
    df_seo = pd.DataFrame(seo_data)

    # --- EXTRACT & TRANSFORM: HTTP Headers ---
    headers_file = cache_dir / "headers.json"
    headers_data = {"Header": [], "Value": []}
    if headers_file.exists():
        try:
            with open(headers_file, 'r', encoding='utf-8') as f:
                h_json = json.load(f)
                actual_headers = h_json.get("headers", {})
                for k, v in actual_headers.items():
                    headers_data["Header"].append(str(k).title())
                    headers_data["Value"].append(str(v))
        except Exception as e:
            print(f"⚠️ Warning: Could not parse headers.json: {e}")
    df_headers = pd.DataFrame(headers_data)

    # --- EXTRACT & TRANSFORM: The ASCII Lenses ---
    ascii_artifacts = {
        'A11y Intent Tree': 'accessibility_tree_summary.txt',
        'Tree Source': 'source_dom_hierarchy.txt',
        'Tree Hydrated': 'hydrated_dom_hierarchy.txt',
        'Tree Diff': 'diff_hierarchy.txt',
        'Simple Source HTML': 'simple_source_html.html',
        'Simple Hydrated DOM': 'simple_hydrated_dom.html',
        'Simple DOM Diff': 'diff_simple_dom.txt'
    }
    
    ascii_dfs = {}
    for sheet_name, filename in ascii_artifacts.items():
        file_path = cache_dir / filename
        if file_path.exists():
            # Read line by line, expanding tabs to hard spaces to prevent Excel from collapsing them
            lines = [line.expandtabs(4) for line in file_path.read_text(encoding='utf-8').splitlines()]
            if 'Diff' in sheet_name:
                structured_data = []
                for line in lines:
                    diff_type = ''
                    if line.startswith('+'):
                        diff_type = 'Added'
                    elif line.startswith('-'):
                        diff_type = 'Removed'
                    elif line.startswith('@@'):
                        diff_type = 'Meta'
                        
                    # Prefix with a space to neutralize Excel formula execution
                    safe_line = f" {line}" if line.startswith(('+', '-', '@', '=')) else line
                    structured_data.append({"Diff Type": diff_type, "Terminal Output": safe_line})
                ascii_dfs[sheet_name] = pd.DataFrame(structured_data)
            else:
                ascii_dfs[sheet_name] = pd.DataFrame({"Terminal Output": lines})

    # --- LOAD: Excel Deliverable ---
    # Isolate the deliverable directory per domain to prevent client name leakage
    domain_slug = domain.replace('.', '_')
    deliverables_dir = wand.paths.deliverables / job / domain_slug
    deliverables_dir.mkdir(parents=True, exist_ok=True)
    xl_filename = f"{domain_slug}_Technical_Baseline.xlsx"
    xl_file = deliverables_dir / xl_filename

    with pd.ExcelWriter(xl_file, engine="xlsxwriter") as writer:
        workbook = writer.book
        
        # Format Definitions
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1, 'align': 'left'})
        wrap_fmt = workbook.add_format({'text_wrap': True, 'valign': 'top'})
        
        # Terminal Formatting
        mono_fmt = workbook.add_format({'font_name': 'Courier New', 'font_size': 9, 'text_wrap': False, 'valign': 'top'})
        add_fmt = workbook.add_format({'font_name': 'Courier New', 'font_size': 9, 'font_color': '#008000', 'bg_color': '#e6ffec'}) # Green
        rem_fmt = workbook.add_format({'font_name': 'Courier New', 'font_size': 9, 'font_color': '#cc0000', 'bg_color': '#ffe6e6'}) # Red
        meta_fmt = workbook.add_format({'font_name': 'Courier New', 'font_size': 9, 'font_color': '#808080'}) # Grey

        # 1. Write Standard Data Tabs
        for sheet_name, df_sheet in [('SEO Metadata', df_seo), ('HTTP Headers', df_headers)]:
            if not df_sheet.empty:
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                ws.set_tab_color('#4F81BD') # Professional Blue
                ws.set_column(0, 0, 30, wrap_fmt)
                ws.set_column(1, 1, 80, wrap_fmt)
                for col_num, value in enumerate(df_sheet.columns.values):
                    ws.write(0, col_num, value, header_fmt)

        # 2. Write Terminal ASCII Tabs
        # 2. Write Terminal ASCII Tabs
        for sheet_name, df_sheet in ascii_dfs.items():
            if not df_sheet.empty:
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                
                # The Tab Color Flex
                if 'A11y' in sheet_name:
                    ws.set_tab_color('#8064A2') # Phantom Purple
                elif 'Tree' in sheet_name:
                    ws.set_tab_color('#9BBB59') # Structural Green
                elif 'Simple' in sheet_name:
                    ws.set_tab_color('#F79646') # Semantic Orange
                
                if 'Diff' in sheet_name:
                    # Hide the 'Diff Type' column, make 'Terminal Output' massive
                    ws.set_column(0, 0, 10, mono_fmt, {'hidden': True})
                    ws.set_column(1, 1, 180, mono_fmt)
                    ws.write(0, 0, "Diff Type", header_fmt)
                    ws.write(0, 1, "Terminal Output", header_fmt)
                    max_row = len(df_sheet) + 1
                    rng = f'A2:B{max_row}'
                    
                    # Conditional formatting perfectly anchored to the hidden helper column
                    ws.conditional_format(rng, {'type': 'formula', 'criteria': '=$A2="Added"', 'format': add_fmt})
                    ws.conditional_format(rng, {'type': 'formula', 'criteria': '=$A2="Removed"', 'format': rem_fmt})
                    ws.conditional_format(rng, {'type': 'formula', 'criteria': '=$A2="Meta"', 'format': meta_fmt})
                else:
                    # Standard ASCII tabs just have 'Terminal Output' (A)
                    ws.set_column(0, 0, 180, mono_fmt)
                    ws.write(0, 0, "Terminal Output", header_fmt)

    # Egress Button
    button = widgets.Button(description=f"📂 Open Deliverables Folder", tooltip=f"Open {deliverables_dir.resolve()}", button_style='success')
    button.on_click(lambda b: wand.open_folder(str(deliverables_dir)))

    return df_seo, df_headers, button, xl_file


def package_optics_to_excel(job: str, target_url: str, ai_assessment: str):
    """
    Packages the high-signal LLM Optics into a beautifully formatted Excel deliverable.
    """
    import pandas as pd
    import re
    import yaml
    from tools.scraper_tools import get_safe_path_component
    from pipulate import wand
    import ipywidgets as widgets

    domain, slug = get_safe_path_component(target_url)
    cache_dir = wand.paths.browser_cache / domain / slug

    # 1. Parse SEO Metadata (The Signal)
    seo_file = cache_dir / "seo.md"
    seo_data = {"Metric": [], "Value": []}
    
    if seo_file.exists():
        content = seo_file.read_text(encoding='utf-8')
        # Extract YAML frontmatter
        match = re.search(r'^---\n(.*?)\n---', content, re.DOTALL)
        if match:
            try:
                frontmatter = yaml.safe_load(match.group(1))
                for k, v in frontmatter.items():
                    seo_data["Metric"].append(str(k).replace('_', ' ').title())
                    seo_data["Value"].append(str(v))
            except Exception as e:
                print(f"⚠️ Warning: Could not parse SEO frontmatter: {e}")

    df_seo = pd.DataFrame(seo_data)

    # 2. Prepare AI Assessment
    df_ai = pd.DataFrame({
        "Intelligence Layer": ["Local Edge AI (Chip O'Theseus)"], 
        "Semantic Assessment": [ai_assessment]
    })

    # 3. Write to Excel with High-End Formatting
    deliverables_dir = wand.paths.deliverables / job
    deliverables_dir.mkdir(parents=True, exist_ok=True)
    xl_filename = f"{domain.replace('.', '_')}_Optics_Baseline.xlsx"
    xl_file = deliverables_dir / xl_filename

    with pd.ExcelWriter(xl_file, engine="xlsxwriter") as writer:
        df_ai.to_excel(writer, sheet_name='AI Assessment', index=False)
        if not df_seo.empty:
            df_seo.to_excel(writer, sheet_name='SEO Metadata', index=False)

        workbook = writer.book
        header_fmt = workbook.add_format({
            'bold': True, 
            'bg_color': '#D9E1F2', 
            'border': 1, 
            'align': 'center'
        })
        wrap_fmt = workbook.add_format({'text_wrap': True, 'valign': 'top'})

        # Format AI Sheet
        ws_ai = writer.sheets['AI Assessment']
        ws_ai.set_column(0, 0, 30, wrap_fmt)
        ws_ai.set_column(1, 1, 100, wrap_fmt)
        for col_num, value in enumerate(df_ai.columns.values):
            ws_ai.write(0, col_num, value, header_fmt)

        # Format SEO Sheet
        if not df_seo.empty:
            ws_seo = writer.sheets['SEO Metadata']
            ws_seo.set_column(0, 0, 25, wrap_fmt)
            ws_seo.set_column(1, 1, 80, wrap_fmt)
            for col_num, value in enumerate(df_seo.columns.values):
                ws_seo.write(0, col_num, value, header_fmt)

    # 4. Generate Egress Button
    button = widgets.Button(
        description=f"📂 Open Deliverables Folder",
        tooltip=f"Open {deliverables_dir.resolve()}",
        button_style='success'
    )
    
    def on_click(b):
        wand.open_folder(str(deliverables_dir))
        
    button.on_click(on_click)

    return button, xl_file


def render_copy_button(prompt_text: str):
    """Renders an HTML/JS button to copy text to the OS clipboard from Jupyter."""
    from IPython.display import HTML
    import base64

    # Base64 encode to safely pass multi-line Python strings into JS
    b64_prompt = base64.b64encode(prompt_text.encode('utf-8')).decode('utf-8')
    
    button_html = f"""
    <button onclick="
        const ta = document.createElement('textarea');
        ta.value = decodeURIComponent(escape(window.atob('{b64_prompt}')));
        document.body.appendChild(ta);
        ta.select();
        document.execCommand('copy');
        document.body.removeChild(ta);
        this.innerText = '✅ Copied to OS Clipboard! Paste into Gemini/Claude.';
        this.style.backgroundColor = '#28a745';
        this.style.color = 'white';
        this.style.borderColor = '#28a745';
    " style="padding: 12px 24px; font-size: 16px; border-radius: 6px; cursor: pointer; border: 1px solid #ccc; font-weight: bold; margin-top: 15px;">
    📋 Copy 'JavaScript Gap' Prompt for Cloud AI
    </button>
    """
    return HTML(button_html)


def conduct_local_assessment(job_id: str, target_url: str, local_model_id: str):
    """
    Orchestrates the local AI assessment, displays the result to the notebook,
    and idempotently injects the findings into the Excel deliverable.
    """
    from IPython.display import display
    from pathlib import Path
    from pipulate import wand

    # 1. Prepare the AI directives
    system_prompt, user_prompt = build_local_optics_prompt(target_url)

    wand.speak(f"Channeling local intent through {local_model_id} to deduce the brand and target keyword...")

    # 2. Execute the local prompt (Kept front-and-center for the user to see)
    ai_assessment = wand.prompt(
        prompt_text=user_prompt, 
        model_name=local_model_id, 
        system_prompt=system_prompt
    )

    print(f"\n🤖 Chip O'Theseus ({local_model_id}):\n")
    print(ai_assessment)
    print("\n" + "-"*40 + "\n")

    # 3. Idempotent Deliverable Injection
    wand.speak("Injecting AI insights directly into your technical baseline workbook.")

    xl_file_path_str = wand.get(job_id, "baseline_excel_path")

    if xl_file_path_str and Path(xl_file_path_str).exists():
        xl_file = Path(xl_file_path_str)
        button, xl_file = append_ai_keyword_assessment(
            job_id, xl_file, ai_assessment, local_model_id, target_url
        )
        display(button)
        print(f"💾 Optics Baseline Augmented: {xl_file.name}")
    else:
        print("⚠️ Technical Baseline Excel file not found. Did you run the Pandas cell?")

    wand.speak(
        "Deliverable upgraded. We have successfully fused raw browser automation "
        "with local generative intelligence. You may open the folder to inspect the result. "
        "Next, we prepare for the Cloud AI handoff."
    )


def factory_reset_credentials():
    """
    Renders a two-step IPyWidget confirmation to wipe API keys from the .env vault
    and the active environment, allowing the onboarding workflow to be tested from scratch.
    """
    import ipywidgets as widgets
    from IPython.display import display, clear_output
    import os
    from dotenv import set_key
    from pipulate import wand
    
    button = widgets.Button(
        description="⚠️ Factory Reset .env Vault",
        button_style="danger",
        icon="trash"
    )
    out = widgets.Output()

    def on_click(b):
        with out:
            clear_output()
            wand.speak("Are you sure? This will wipe your cloud keys.")
            confirm = widgets.Button(description="Yes, Wipe It", button_style="danger")
            cancel = widgets.Button(description="Cancel", button_style="success")
            
            def on_confirm(cb):
                with out:
                    clear_output()
                    env_path = wand.paths.root / ".env"
                    keys_to_nuke = [
                        "OPENAI_API_KEY", 
                        "ANTHROPIC_API_KEY", 
                        "GEMINI_API_KEY",
                        "GOOGLE_API_KEY",
                        "BOTIFY_API_TOKEN"
                    ]
                    for k in keys_to_nuke:
                        if k in os.environ:
                            del os.environ[k]
                    # Ruthless file scrubber: physically purge the lines
                    if env_path.exists():
                        lines = env_path.read_text(encoding='utf-8').splitlines()
                        clean_lines = [line for line in lines if not any(line.startswith(f"{k}=") for k in keys_to_nuke)]
                        if clean_lines:
                            env_path.write_text('\n'.join(clean_lines) + '\n', encoding='utf-8')
                        else:
                            env_path.unlink()  # Nuke the whole file if it's empty
                    
                    wand.speak("Vault wiped. Restart the kernel to complete the amnesia.")
                    print("✅ Credentials cleared. Please restart the kernel (Esc, 0, 0) to start over.")

            def on_cancel(cb):
                with out:
                    clear_output()
                    wand.speak("Crisis averted.")
                    print("🛡️ Crisis averted. Vault remains intact.")

            confirm.on_click(on_confirm)
            cancel.on_click(on_cancel)
            display(widgets.HBox([confirm, cancel]))

    button.on_click(on_click)
    display(widgets.VBox([button, out]))

def render_persona_selector(job_id: str = "onboarding_job"):
    """
    Renders the IPyWidget to select the AI auditor persona.
    Handles persistent state writing and triggers the downstream compulsion.
    """
    import ipywidgets as widgets
    from IPython.display import display, clear_output
    from pipulate import wand

    # 1. Check if they already made a choice previously (defaulting to the suit)
    existing_choice = wand.get(job_id, "auditor_persona") or "enterprise"

    # 2. Build the visual selection widget
    persona_widget = widgets.RadioButtons(
        options=[
            ('👔 The Enterprise Consultant (Strict, analytical, buttoned-up)', 'enterprise'),
            ('🎭 Statler & Waldorf (Ruthless heckling from the balcony)', 'muppets')
        ],
        value=existing_choice,
        layout={'width': 'max-content'}
    )

    submit_btn = widgets.Button(
        description="Lock in Persona", 
        button_style='primary',
        icon='check'
    )

    out = widgets.Output()

    def on_submit(b):
        with out:
            clear_output()
            selected = persona_widget.value
            
            # 3. Save the choice to persistent memory
            wand.set(job_id, "auditor_persona", selected)
            
            # 4. Give contextual audio feedback based on their choice
            if selected == 'muppets':
                wand.speak("Excellent choice. Prepare to be insulted.")
            else:
                wand.speak("Very well. We will keep this strictly professional.")
                
            submit_btn.description = "Persona Locked"
            submit_btn.button_style = 'success'
            
            # 5. Fire the compulsion to advance
            wand.speak("You may now run the next cell.", emoji="✅") # REPLACED

    submit_btn.on_click(on_submit)

    display(widgets.VBox([persona_widget, submit_btn, out]))


def prepare_prompt_draft(job_id: str, recovered_url: str, local_model: str):
    """
    Asks the local AI to look at the diff artifacts and draft the initial 
    prompt for the Cloud AI.
    """
    from tools.scraper_tools import get_safe_path_component
    
    # 1. Locate the artifacts
    domain, slug = get_safe_path_component(recovered_url)
    cache_base = wand.paths.browser_cache / domain / slug
    diff_file = cache_base / "diff_hierarchy.txt"
    
    diff_context = "No diff data available."
    if diff_file.exists():
        # Grab the middle 2000 chars of the diff where the interesting stuff usually is
        content = diff_file.read_text()
        diff_context = content[:2000]

    # 2. Determine Persona Scripting
    persona = wand.get(job_id, "auditor_persona") or "enterprise"
    
    if persona == "muppets":
        system_msg = "You are Statler and Waldorf. You are heckling a web developer's messy JavaScript-heavy site."
        task_msg = "Write a snarky, insulting prompt for a Cloud AI to analyze this DOM diff. Be mean but technical."
    else:
        system_msg = "You are an elite Enterprise SEO Consultant. You are professional and surgical."
        task_msg = "Draft a formal technical audit prompt for a frontier model to analyze this DOM diff."

    prompt_to_local = f"""
    {task_msg}
    
    URL: {recovered_url}
    DOM DIFF SNIPPET:
    {diff_context}
    
    Generate the 'Instructions' portion of a prompt. Do not include the data envelope.
    """

    # 3. Call Local AI via Wand
    draft = wand.prompt(prompt_to_local, model_name=local_model, system_prompt=system_msg)
    
    # 4. Save to Disk and store Pointer in Wand Memory
    job_dir = wand.paths.data / "jobs" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    prompt_file = job_dir / "cloud_prompt.md"
    prompt_file.write_text(draft.strip(), encoding='utf-8')
    wand.set(job_id, "cloud_prompt_path", str(prompt_file))
    return draft


def render_prompt_workbench(job_id: str, recovered_url: str):
    """
    Renders an editable textarea with a layout that doesn't 
    clobber subsequent cell output.
    """
    import ipywidgets as widgets
    from IPython.display import display, HTML
    from tools.scraper_tools import get_safe_path_component
    from pipulate import wand

    # 1. Gather context for the 'Verification Links'
    domain, slug = get_safe_path_component(recovered_url)
    cache_base = wand.paths.browser_cache / domain / slug

    try:
        rel_dir = cache_base.relative_to(wand.paths.root)
        rel_dir_str = str(rel_dir).replace('\\', '/')
    except ValueError:
        rel_dir_str = str(cache_base).replace('\\', '/')

    # 2. Fetch drafted prompt from Disk (via Pointer)
    prompt_path_str = wand.get(job_id, "cloud_prompt_path")
    existing_prompt = "Drafting..."
    if prompt_path_str and Path(prompt_path_str).exists():
        try:
            existing_prompt = Path(prompt_path_str).read_text(encoding='utf-8')
        except Exception as e:
            existing_prompt = f"Error reading prompt file: {e}"

    # 3. Build UI components
    prompt_area = widgets.Textarea(
        value=existing_prompt,
        layout=widgets.Layout(width='98%', height='400px')
    )

    save_btn = widgets.Button(
        description="💾 Save & Finalize Prompt",
        button_style='success',
        layout=widgets.Layout(width='250px')
    )

    # Display the instructions above the textarea so the user can 'verify' before saving
    link_html = HTML(f"""
        <div style="margin-bottom: 10px; padding: 10px; border: 1px solid #ccc; border-radius: 5px;">
            <strong>🔍 Verification Step:</strong><br>
            Use the Jupyter file browser on the left to navigate to:<br>
            <code>{rel_dir_str}/</code><br><br>
            Double-click <strong>diff_hierarchy.html</strong> and <strong>diff_boxes.html</strong> to view the visual evidence before saving.
        </div>
    """)

    def on_save(b):
        if prompt_path_str:
            Path(prompt_path_str).write_text(prompt_area.value, encoding='utf-8')
        save_btn.description = "✅ Saved to Disk"
        save_btn.button_style = ''
        wand.speak("Instructions locked. Ready for the next turn.")
        # Trigger the visual compulsion below the widget
        with out:
            wand.speak("You may now run the next cell.", emoji="✅") # REPLACED

    save_btn.on_click(on_save)
    
    # 4. Use an Output widget to contain everything
    out = widgets.Output()
    display(link_html, prompt_area, save_btn, out)


def render_cloud_handoff(job_id: str, recovered_url: str):
    """
    Retrieves the user-polished prompt, compiles the DOM diff data JIT,
    and renders a Bifurcated Egress (Copy Button + Paste Bin) for loose coupling.
    """
    import ipywidgets as widgets
    from tools.scraper_tools import get_safe_path_component
    from IPython.display import HTML

    # 1. Retrieve the polished instructions
    instructions = wand.get(job_id, "cloud_ai_prompt")
    if not instructions:
        return widgets.HTML("<p style='color:var(--pico-color-red-500);'>⚠️ No instructions found. Did you click 'Save'?</p>"), ""

    # 2. Store the Absolute Path Reference (The Pointer)
    domain, slug = get_safe_path_component(recovered_url)
    cache_base = wand.paths.browser_cache / domain / slug

    diff_path = cache_base / "diff_simple_dom.txt"
    wand.set(job_id, "optics_diff_path", str(diff_path))

    # 3. JIT Compile the Final Payload for the UI Copy Button
    final_payload = compile_cloud_payload(job_id, recovered_url)

    # 4. Build the Bifurcated UI (Copy Button + Paste Bin)
    paste_area = widgets.Textarea(
        value=wand.get(job_id, "manual_cloud_response") or "",
        placeholder="Paste the Web UI response here (leave blank to use the formal API)...",
        layout=widgets.Layout(width='98%', height='150px')
    )
    
    def on_change(change):
        wand.set(job_id, "manual_cloud_response", change['new'])
        
    paste_area.observe(on_change, names='value')
    
    # Extract the raw HTML string from the IPython display object
    copy_btn_html = render_copy_button(final_payload).data
    
    ui = widgets.VBox([
        widgets.HTML("<p><b>Option 1: The 'Poor-Man's API' (Free)</b><br>Click below to copy the prompt, paste it into ChatGPT/Claude, and paste the response into the text box below.</p>"),
        widgets.HTML(copy_btn_html),
        widgets.HTML("<br><p><b>Paste Web UI Response Here:</b><br>(If you leave this blank, running the next cell will automatically use your API key)</p>"),
        paste_area
    ])

    return ui, final_payload


def reveal_system_architecture():
    from rich.console import Console
    from rich.panel import Panel
    from rich.text import Text
    
    console = Console()
    lens_art = """
  Idea --> Lens 1   -->   Lens 2  -->  Lens 3  -> Lens 4 -> Lens 5 -> Lens 6

     -----> ,--.
     ---> ,'    `.---------> ,--.
     --> /        \------> ,'    `.-------> ,--.        ,-.
  o  -> /  Linux   \----> /  HTTP  \----> ,'_hx `.--->,'   `.    ,-.
 /|\   (  HARDWARE  )--> ( PROTOCOL )--> ( LINGUA )->( UI/UX )->(APP)->(git)
 / \ -> \  (Nix)   /----> \ (html) /----> `..py ,'--->`.   ,'    `-'
     --> \        /------> `.    ,'-------> `--'        `-'    And so on
     ---> `.    ,'---------> `--'         AI Help
     -----> `--'           AI Help
          AI Help
    """
    
    # We apply specific colors to specific layers of the stack
    styled_art = Text(lens_art)
    styled_art.highlight_regex(r"HARDWARE|Linux|Nix", "bold cyan")
    styled_art.highlight_regex(r"PROTOCOL|http|html", "bold green")
    styled_art.highlight_regex(r"LINGUA|_hx|\.py", "bold yellow")
    styled_art.highlight_regex(r"UI/UX", "bold magenta")
    styled_art.highlight_regex(r"APP|git", "bold blue")
    styled_art.highlight_regex(r"AI Help", "dim white")
    
    console.print(Panel(styled_art, title="[bold orange3]The Pipulate Lens Stack[/]", border_style="cyan"))


def build_local_optics_prompt(target_url: str):
    """Generates the local prompt to extract the target keyword from SEO metadata."""
    from tools.scraper_tools import get_safe_path_component
    from pipulate import wand
    import re
    import yaml
    import pandas as pd

    domain, slug = get_safe_path_component(target_url)
    seo_file = wand.paths.browser_cache / domain / slug / "seo.md"

    seo_context = "No SEO data available."
    if seo_file.exists():
        content = seo_file.read_text(encoding='utf-8')
        match = re.search(r'^---\n(.*?)\n---', content, re.DOTALL)
        if match:
            try:
                frontmatter = yaml.safe_load(match.group(1))
                seo_data = {"Metric": [], "Value": []}
                for k, v in frontmatter.items():
                    seo_data["Metric"].append(str(k).replace('_', ' ').title())
                    seo_data["Value"].append(str(v))
                seo_context = pd.DataFrame(seo_data).to_string(index=False)
            except Exception:
                pass

    local_system_prompt = (
        "You are Chip O'Theseus, an AI running locally on the user's hardware. "
        "You are an expert technical SEO."
    )

    local_prompt = f"""
Analyze this metadata extracted from a webpage:

URL: {target_url}

METADATA:
{seo_context}

Based strictly on this data, identify the Brand Entity and the primary Generic Keyword this page is trying to target.
Respond with exactly four lines:
BRAND: [The brand or company name]
BRAND_RATIONALE: [One sentence explaining why]
KEYWORD: [The primary generic targeted keyword phrase]
KEYWORD_RATIONALE: [One sentence explaining why based on the title/h1 tags]
"""
    return local_system_prompt, local_prompt.strip()

def append_ai_keyword_assessment(job: str, xl_file_path, ai_assessment: str, local_model_id: str, target_url: str):
    """
    Idempotently appends a local AI assessment tab to an existing Excel deliverable,
    with high-end enterprise formatting.
    """
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from pipulate import wand
    from datetime import datetime
    import ipywidgets as widgets
    from pathlib import Path

    # 1. Idempotency Check
    book = openpyxl.load_workbook(xl_file_path)
    if 'AI Keyword Target' in book.sheetnames:
        print("☑️ 'AI Keyword Target' tab already exists in this workbook.")
    else:
        # 2. Deterministic Parsing of the LLM Output (Robust against blank lines)
        lines = [line.strip() for line in ai_assessment.strip().split('\n') if line.strip()]
        
        brand, brand_rat = "Unknown", "Failed to parse."
        keyword, keyword_rat = "Unknown", "Failed to parse."
        
        for line in lines:
            if line.startswith("BRAND:"): brand = line.replace("BRAND:", "").strip()
            elif line.startswith("BRAND_RATIONALE:"): brand_rat = line.replace("BRAND_RATIONALE:", "").strip()
            elif line.startswith("KEYWORD:"): keyword = line.replace("KEYWORD:", "").strip()
            elif line.startswith("KEYWORD_RATIONALE:"): keyword_rat = line.replace("KEYWORD_RATIONALE:", "").strip()
        
        df_ai = pd.DataFrame({
            "Crawled URL": [target_url, target_url],
            "Entity Type": ["Brand", "Generic Keyword"],
            "Predicted Target": [brand, keyword],
            "AI Rationale": [brand_rat, keyword_rat],
            "Model Used": [local_model_id, local_model_id],
            "Timestamp": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")] * 2
        })
        
        # 3. The Safe Load (Writing the new tab)
        with pd.ExcelWriter(xl_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_ai.to_excel(writer, sheet_name='AI Keyword Target', index=False)
            
            # --- THE AESTHETIC UPGRADE ---
            ws = writer.sheets['AI Keyword Target']
            ws.sheet_properties.tabColor = "FFC000"  # AI Gold
            
            max_col = get_column_letter(ws.max_column)
            max_row = ws.max_row
            
            # Format as a slick Excel Table
            tab = Table(displayName="AIKeywordTarget", ref=f"A1:{max_col}{max_row}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)
            
            # Set specific column widths for readability
            widths = {'A': 25, 'B': 18, 'C': 25, 'D': 80, 'E': 18, 'F': 20}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
                
            # Apply distinct formatting for Headers vs Data
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=ws.max_column), start=1):
                for cell in row:
                    if row_idx == 1:
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        cell.alignment = wrap_alignment

        print(f"✅ AI Insights successfully appended to {Path(xl_file_path).name}")
    
    # 4. SOVEREIGN DELIVERY FIX: Open exactly one level deeper
    deliverables_dir = Path(xl_file_path).parent
    button = widgets.Button(description=f"📂 Open Deliverables Folder", tooltip=f"Open {deliverables_dir.resolve()}", button_style='success')
    button.on_click(lambda b: wand.open_folder(str(deliverables_dir)))
    
    return button, xl_file_path


def append_cloud_assessment(job: str, xl_file_path, ai_assessment: str, model_id: str):
    """
    Idempotently appends the Cloud AI JavaScript Gap analysis to the Excel deliverable,
    with high-end enterprise formatting.
    """
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from pipulate import wand
    from datetime import datetime
    import ipywidgets as widgets
    from pathlib import Path

    book = openpyxl.load_workbook(xl_file_path)
    if 'Cloud JS Gap Analysis' in book.sheetnames:
        print("☑️ 'Cloud JS Gap Analysis' tab already exists in this workbook.")
    else:
        df_ai = pd.DataFrame({
            "Intelligence Layer": ["Cloud Frontier Model"],
            "Semantic Assessment": [ai_assessment],
            "Model Used": [model_id],
            "Timestamp": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        })
        
        with pd.ExcelWriter(xl_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_ai.to_excel(writer, sheet_name='Cloud JS Gap Analysis', index=False)
            
            # --- THE AESTHETIC UPGRADE ---
            ws = writer.sheets['Cloud JS Gap Analysis']
            ws.sheet_properties.tabColor = "FF33CC"  # Cloud Pink/Magenta
            
            max_col = get_column_letter(ws.max_column)
            max_row = ws.max_row
            
            tab = Table(displayName="CloudAnalysis", ref=f"A1:{max_col}{max_row}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            widths = {'A': 25, 'B': 100, 'C': 18, 'D': 20}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
                
            # Apply distinct formatting for Headers vs Data
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=ws.max_column), start=1):
                for cell in row:
                    if row_idx == 1:
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        cell.alignment = wrap_alignment

        print(f"✅ Cloud Insights successfully appended to {Path(xl_file_path).name}")
    
    # SOVEREIGN DELIVERY FIX: Open exactly one level deeper
    deliverables_dir = Path(xl_file_path).parent
    button = widgets.Button(description=f"📂 Open Deliverables Folder", tooltip=f"Open {deliverables_dir.resolve()}", button_style='success')
    button.on_click(lambda b: wand.open_folder(str(deliverables_dir)))
    
    return button, Path(xl_file_path)


def compile_cloud_payload(job_id: str, target_url: str) -> str:
    """
    JIT compilation of the Cloud AI prompt.
    Reads the user instructions from the wand and heavy artifacts from the disk.
    """
    from tools.scraper_tools import get_safe_path_component
    from pipulate import wand

    instructions = wand.get(job_id, "cloud_ai_prompt") or "Please analyze the following data."
    prompt_path_str = wand.get(job_id, "cloud_prompt_path")
    if prompt_path_str and Path(prompt_path_str).exists():
        instructions = Path(prompt_path_str).read_text(encoding='utf-8')
    
    # Resolve the pointer
    domain, slug = get_safe_path_component(target_url)
    cache_base = wand.paths.browser_cache / domain / slug

    # The Flexible Artifact Roster (Ordered by Priority)
    artifacts_to_include = [
        ("Semantic Outline", "accessibility_tree_summary.txt"),
        ("Unified Diff Snippet", "diff_simple_dom.txt")
    ]

    payload_parts = [instructions, "\n\n# ARTIFACTS\n"]
    
    # Max payload size for UI clipboard safety (approx 250KB)
    MAX_BYTES = 250 * 1024
    current_bytes = len("".join(payload_parts).encode('utf-8'))

    for label, filename in artifacts_to_include:
        file_path = cache_base / filename
        if not file_path.exists():
            continue

        content = file_path.read_text(encoding='utf-8')
        
        # Apply prompt_foo style delimiters
        header = f"\n--- START: {label} ({filename}) ---\n```text\n"
        footer = "\n```\n--- END ---\n"
        
        overhead_bytes = len((header + footer).encode('utf-8'))
        available_bytes = MAX_BYTES - current_bytes - overhead_bytes
        
        if available_bytes <= 0:
            break  # No room left for more files

        content_bytes = len(content.encode('utf-8'))
        
        if content_bytes > available_bytes:
            # Truncate safely by bytes to respect clipboard limits
            truncated_bytes = content.encode('utf-8')[:available_bytes]
            content = truncated_bytes.decode('utf-8', errors='ignore') + "\n...[TRUNCATED TO FIT CAPACITY]..."
        
        formatted_part = f"{header}{content}{footer}"
        payload_parts.append(formatted_part)
        current_bytes += len(formatted_part.encode('utf-8'))

    final_payload = "".join(payload_parts)

    # Write the fully compiled payload to disk for the fossil record
    job_dir = wand.paths.data / "jobs" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    compiled_file = job_dir / "compiled_payload.md"
    compiled_file.write_text(final_payload, encoding='utf-8')
    wand.set(job_id, "compiled_payload_path", str(compiled_file))
    
    return final_payload


def execute_cloud_analysis(job_id: str, recovered_url: str, active_cloud_model: str):
    """
    Checks for a manual response, falls back to the API with exponential backoff, 
    renders the output via Rich, and injects the final assessment into the Excel baseline.
    """
    from IPython.display import display
    from pathlib import Path
    import time
    from pipulate import wand

    # 1. Check the manual paste bin from the previous step
    manual_response = wand.get(job_id, "manual_cloud_response")
    final_analysis = ""
    active_model_used = "None"

    if manual_response and manual_response.strip():
        wand.speak("Manual response detected in the paste bin. Bypassing the metered API.")
        final_analysis = manual_response
        active_model_used = "Manual Web UI Paste"
    else:
        wand.speak(f"No manual response detected. Engaging formal API via {active_cloud_model}...")
        payload = compile_cloud_payload(job_id, recovered_url)
        
        if payload:
            active_model_used = active_cloud_model
            max_retries = 3
            base_wait = 5
            
            # Exponential Backoff Loop
            for attempt in range(max_retries):
                try:
                    final_analysis = wand.prompt(prompt_text=payload, model_name=active_cloud_model)
                    break  # Success, exit retry loop
                except Exception as e:
                    error_msg = str(e).lower()
                    if any(trigger in error_msg for trigger in ["429", "500", "503", "high demand", "quota"]):
                        if attempt < max_retries - 1:
                            wait_time = base_wait * (2 ** attempt)  # 5s, 10s, 20s
                            print(f"⚠️ {active_cloud_model} is experiencing network friction. Retrying in {wait_time} seconds...")
                            time.sleep(wait_time)
                        else:
                            final_analysis = f"❌ API failed after {max_retries} attempts: {e}"
                            print(final_analysis)
                    else:
                        final_analysis = f"❌ Execution Error: {e}"
                        print(final_analysis)
                        break
        else:
            final_analysis = "Error: Payload missing. Did you run the previous steps?"

    # 2. Use Rich for beautiful output if the terminal supports it
    try:
        from rich.console import Console
        from rich.panel import Panel
        from rich.markdown import Markdown
        console = Console()
        console.print(Panel(Markdown(final_analysis), title=f"[bold cyan]☁️ Cloud AI Analysis ({active_model_used})[/]", border_style="blue"))
    except ImportError:
        print(f"\n☁️ Cloud AI Analysis ({active_model_used}):\n{'-'*40}\n{final_analysis}\n{'-'*40}\n")

    # 3. The Final Stamp: Idempotent Deliverable Injection
    wand.speak("The audit is complete. I am injecting the Cloud AI insights into your technical baseline workbook.")
    xl_file_path_str = wand.get(job_id, "baseline_excel_path")

    if xl_file_path_str and Path(xl_file_path_str).exists():
        button, xl_file = append_cloud_assessment(job_id, xl_file_path_str, final_analysis, active_model_used)
        display(button)
        print(f"💾 Optics Baseline Augmented: {xl_file.name}")
    else:
        print("⚠️ Technical Baseline Excel file not found. Did you run the Pandas cell?")
