# assets/nbs/imports/core_sauce.py
import asyncio
import sys
import pandas as pd
from pathlib import Path
from loguru import logger
from pipulate import wand
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import nbformat

# In assets/nbs/imports/core_sauce.py

async def universal_scrape(job, urls, headless=True, delay_range=(5, 10), override_cache=False):
    """The unified acquisition engine for all tier-2 tools."""
    extracted_data = []
    for i, url in enumerate(urls):
        current_delay = delay_range if i > 0 else None
        
        # Pass override_cache down to the wand
        result = await wand.scrape(url=url, headless=headless, delay_range=current_delay, override_cache=override_cache)
        
        is_cached = result.get("cached", False)
        status = "✅ Cached" if is_cached else "👁️ Scraped"
        print(f"  -> {status} [{i+1}/{len(urls)}] {url}")

        dom_path = result.get("looking_at_files", {}).get("hydrated_dom")
        if dom_path and Path(dom_path).exists():
            with open(dom_path, 'r', encoding='utf-8') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                extracted_data.append({
                    'url': url,
                    'title': soup.title.string.strip() if soup.title else "No Title",
                    'h1s': [h1.get_text(strip=True) for h1 in soup.find_all('h1')],
                    'hydrated_dom_path': str(dom_path),
                    'cached': is_cached,                     # <-- New enriched state
                    'success': result.get('success', False), # <-- New enriched state
                    'error': result.get('error', '')         # <-- New enriched state
                })
        else:
            # Handle failure states gracefully in the returned list
            extracted_data.append({
                'url': url,
                'cached': is_cached,
                'success': False,
                'error': result.get('error', 'DOM file not found')
            })
            
    wand.set(job, "extracted_data", extracted_data)
    return extracted_data

async def generate_optics_batch(job, verbose=False):
    """Fires the LLM Optics engine for every URL in the job state."""
    extracted_data = wand.get(job, "extracted_data", [])
    # Correct path resolution for the Forever Machine
    script_path = (Path(__file__).resolve().parent.parent.parent / "tools" / "llm_optics.py")
    
    tasks = []
    for item in extracted_data:
        dom_path = item.get('hydrated_dom_path')
        if dom_path and Path(dom_path).exists():
            tasks.append(_run_optics_subprocess(sys.executable, script_path, dom_path))
            
    if tasks:
        logger.info(f"🚀 Launching {len(tasks)} LLM Optics subprocesses...")
        await asyncio.gather(*tasks)

async def _run_optics_subprocess(py_exe, script, dom):
    proc = await asyncio.create_subprocess_exec(
        py_exe, str(script), str(dom),
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE
    )
    await proc.communicate()

def format_excel_pro(job, df, sheet_name="Analysis"):
    """The master 'Painterly Pass' for Excel output."""
    output_dir = Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)
    xl_file = output_dir / f"{job}_output.xlsx"
    
    with pd.ExcelWriter(xl_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        tab = Table(displayName="DataTable", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 30
            for cell in col: cell.alignment = Alignment(wrap_text=True, vertical='top')
    print(f"✅ Deliverable saved to: {xl_file}")
    return xl_file


def extract_tagged_cell(notebook_filename: str, tag: str) -> str:
    """Topological search across the workspace to extract a tagged notebook cell."""
    if not notebook_filename.endswith('.ipynb'):
        notebook_filename += '.ipynb'
        
    # Cast a recursive net across the entire Notebooks/ base directory
    nb_path = next(wand.paths.base.rglob(notebook_filename), None)
    
    if not nb_path:
        logger.error(f"❌ Could not find {notebook_filename} anywhere in {wand.paths.base}")
        return ""
        
    try:
        with open(nb_path, 'r', encoding='utf-8') as f:
            nb = nbformat.read(f, as_version=4)
            for cell in nb.cells:
                if tag in cell.metadata.get("tags", []):
                    return cell.source
    except Exception as e:
        logger.error(f"❌ Error reading {nb_path}: {e}")
    
    return ""

def get_urls_from_notebook(notebook_filename: str) -> list:
    """Extracts clean URLs from the 'url-list-input' cell."""
    source = extract_tagged_cell(notebook_filename, "url-list-input")
    return [line.split('#')[0].strip() for line in source.splitlines() 
            if line.strip() and not line.strip().startswith('#')]

def get_prompt_from_notebook(notebook_filename: str) -> str:
    """Extracts the system prompt from the 'prompt-input' cell."""
    return extract_tagged_cell(notebook_filename, "prompt-input")

